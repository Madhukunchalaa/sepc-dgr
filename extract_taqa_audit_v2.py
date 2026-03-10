import openpyxl
import datetime

file_path = 'TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
target_dates = [
    datetime.datetime(2025, 7, 15),
    datetime.datetime(2025, 11, 23),
    datetime.datetime(2025, 4, 9)
]

def get_excel_data(dates):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['DGR']
    
    # Dates are in row 9 (1-indexed) according to previous logs, but let's find them
    date_row = 9
    date_cols = {}
    for col in range(4, 400): # From D onwards
        val = sheet.cell(row=date_row, column=col).value
        if isinstance(val, (datetime.datetime, datetime.date)):
            # Normalize to date for comparison
            d = val if isinstance(val, datetime.date) else val.date()
            for td in dates:
                if d == td.date():
                    date_cols[d.strftime('%Y-%m-%d')] = col
                    break
    
    print(f"Found dates in columns: {date_cols}")
    
    # Rows 12 to 150 cover most KPIs
    results = {}
    for d_str, col_idx in date_cols.items():
        day_data = []
        for r in range(12, 160):
            sn = sheet.cell(row=r, column=1).value
            particulars = sheet.cell(row=r, column=2).value
            uom = sheet.cell(row=r, column=3).value
            daily = sheet.cell(row=r, column=col_idx).value
            
            if particulars:
                day_data.append({
                    'sn': str(sn) if sn else '',
                    'particulars': str(particulars).strip(),
                    'uom': str(uom).strip() if uom else '',
                    'daily': daily
                })
        results[d_str] = day_data
        
    return results

results = get_excel_data(target_dates)
if results:
    import json
    with open('excel_audit_values.json', 'w') as f:
        json.dump(results, f, default=str, indent=2)
    print("Saved audit values to excel_audit_values.json")
    for date, data in results.items():
        print(f"\n--- {date} (showing first 5) ---")
        for item in data[:5]:
            print(f"[{item['sn']}] {item['particulars']} ({item['uom']}): {item['daily']}")
