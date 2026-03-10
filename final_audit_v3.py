from openpyxl import load_workbook
import datetime

def audit_v3():
    excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb['Ops Input']

    # 1. Identify rows by searching all first 5 columns for keywords
    row_map = {}
    for r in range(1, 500):
        found_text = ""
        for c in range(1, 6):
            val = ws.cell(row=r, column=c).value
            if val: found_text += str(val) + " "
        
        if "Gross Generation" in found_text and "Main Meter" in found_text:
            row_map['gen'] = r
        if "PAF" in found_text:
            row_map['paf'] = r
        if "PLF" in found_text:
            row_map['plf'] = r
        if "Auxiliary Power Consumption" in found_text or "Aux. Power Consumption" in found_text:
            row_map['apc'] = r

    print(f"Row Map: {row_map}")

    # 2. Identify columns
    # Based on earlier check, Row 1 has the dates starting from Col 5?
    target_date = datetime.datetime(2026, 2, 8).date()
    mtd_start = datetime.datetime(2026, 2, 1).date()
    ytd_start = datetime.datetime(2025, 4, 1).date()

    cols = {'target': None, 'mtd': None, 'ytd': None}
    for col in range(1, 400):
        val = ws.cell(row=1, column=col).value
        # If not in row 1, check row 4
        if not isinstance(val, datetime.datetime):
            val = ws.cell(row=4, column=col).value
            
        if isinstance(val, datetime.datetime):
            d = val.date()
            if d == target_date: cols['target'] = col
            if d == mtd_start: cols['mtd'] = col
            if d == ytd_start: cols['ytd'] = col

    print(f"Column Map: {cols}")

    if cols['target'] and cols['ytd']:
        def calc_sums(row_key, divider=1, is_percentage=False):
            if row_key not in row_map: return 0, 0
            row = row_map[row_key]
            # MTD sum (start of month to target - 1?)
            # User says Feb 8 DAILY is 0. 
            # Usually DGR for Feb 8 includes values cumulative up to Feb 8.
            # But let's check Feb 7.
            
            mtd_sum = 0
            for c in range(cols['mtd'], cols['target'] + 1):
                v = ws.cell(row=row, column=c).value
                if v: mtd_sum += float(v)
            
            ytd_sum = 0
            for c in range(cols['ytd'], cols['target'] + 1):
                v = ws.cell(row=row, column=c).value
                if v: ytd_sum += float(v)
                
            return mtd_sum / divider, ytd_sum / divider

        if 'gen' in row_map:
            m, y = calc_sums('gen', divider=1000)
            print(f"\nGross Generation (MU): MTD={m:.4f}, YTD={y:.4f}")
            print(f"User App Values:    MTD=12.9940, YTD=1304.1757")

audit_v3()
