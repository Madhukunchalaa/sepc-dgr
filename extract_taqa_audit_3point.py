import openpyxl
import datetime
import json

file_path = 'TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'

def extract_dgr_active():
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # --- 1. Active DGR Sheet ---
    sheet_dgr = wb['DGR']
    report_date = sheet_dgr.cell(row=3, column=5).value
    if isinstance(report_date, datetime.datetime):
        report_date_str = report_date.strftime('%Y-%m-%d')
    else:
        report_date_str = str(report_date)
        
    print(f"DGR Sheet Active Date: {report_date_str}")
    
    dgr_values = {}
    for r in range(4, 200):
        # We need to find rows with SN, Particulars, Daily, MTD, YTD
        # In DGR sheet, Row 11 has headers? 
        # Column 1=SN, 2=Particulars, 3=UoM, 4=?, 5=Daily, 6=MTD, 7=YTD?
        sn = str(sheet_dgr.cell(row=r, column=1).value or '').strip()
        particulars = str(sheet_dgr.cell(row=r, column=2).value or '').strip()
        uom = str(sheet_dgr.cell(row=r, column=3).value or '').strip()
        daily = sheet_dgr.cell(row=r, column=5).value
        mtd = sheet_dgr.cell(row=r, column=6).value
        ytd = sheet_dgr.cell(row=r, column=7).value
        
        if particulars and particulars != 'None' and particulars != 'Particulars':
             dgr_values[particulars] = {
                 'sn': sn, 'uom': uom, 'daily': daily, 'mtd': mtd, 'ytd': ytd
             }
             
    # --- 2. DGR (WU) Sheet (2025-04-01) ---
    sheet_wu = wb['DGR (WU)']
    wu_date = '2025-04-01'
    wu_col = 8 # Found earlier
    wu_values = {}
    for r in range(4, 80):
        particulars = str(sheet_wu.cell(row=r, column=3).value or '').strip()
        uom = str(sheet_wu.cell(row=r, column=4).value or '').strip()
        val = sheet_wu.cell(row=r, column=wu_col).value
        if particulars and particulars != 'None':
            wu_values[particulars] = {'uom': uom, 'val': val}

    # --- 3. Raw Data for 2025-08-15 from Ops Input ---
    sheet_ops = wb['Ops Input']
    aug15_col = 115 + 26 # Approx, let's find it exactly
    aug15_date = datetime.date(2025, 8, 15)
    found_col = -1
    for c in range(4, 500):
        val = sheet_ops.cell(row=1, column=c).value
        if isinstance(val, (datetime.datetime, datetime.date)):
             if (val.date() if isinstance(val, datetime.datetime) else val) == aug15_date:
                 found_col = c
                 break
    
    ops_raw = {}
    if found_col != -1:
        for r in range(4, 180):
             particulars = str(sheet_ops.cell(row=r, column=2).value or '').strip()
             val = sheet_ops.cell(row=r, column=found_col).value
             if particulars and particulars != 'None':
                 ops_raw[particulars] = val
                 
    audit_data = {
        'report_date': report_date_str,
        'dgr_active': dgr_values,
        'wu_2025_04_01': wu_values,
        'ops_raw_2025_08_15': ops_raw
    }
    
    with open('target_audit_data.json', 'w') as f:
        json.dump(audit_data, f, default=str, indent=2)
    print("Exported target_audit_data.json")

extract_dgr_active()
