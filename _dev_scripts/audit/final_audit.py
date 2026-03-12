from openpyxl import load_workbook
import datetime

def audit():
    excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb['Ops Input']

    # 1. Identify rows
    rows = {}
    for r in range(1, 400):
        lbl = str(ws.cell(row=r, column=3).value)
        if "Gross Generation" in lbl and "Main Meter" in lbl: rows['gen'] = r
        if "PAF" in lbl: rows['paf'] = r
        if "PLF" in lbl: rows['plf'] = r
        # Add more if needed

    print(f"Row Map: {rows}")

    # 2. Identify columns
    target_date = datetime.datetime(2026, 2, 8).date()
    mtd_start = datetime.datetime(2026, 2, 1).date()
    ytd_start = datetime.datetime(2025, 4, 1).date()

    cols = {'target': None, 'mtd': None, 'ytd': None}

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if isinstance(val, datetime.datetime):
            d = val.date()
            if d == target_date: cols['target'] = col
            if d == mtd_start: cols['mtd'] = col
            if d == ytd_start: cols['ytd'] = col

    print(f"Column Map: {cols}")

    # 3. Calculate Sums for Gross Generation
    if rows.get('gen') and cols['target'] and cols['ytd']:
        gen_row = rows['gen']
        
        # YTD Sum (April 1 to target_date - 1?)
        # User snippet for Feb 8 says DAILY 0, MTD 12.9940.
        # This means MTD sum up to Feb 7.
        
        def safe_sum(row, start_col, end_col):
            total = 0
            for c in range(start_col, end_col + 1):
                v = ws.cell(row=row, column=c).value
                if v: total += float(v)
            return total

        ytd_val = safe_sum(gen_row, cols['ytd'], cols['target']) / 1000 # MWhr to MU
        mtd_val = safe_sum(gen_row, cols['mtd'], cols['target']) / 1000 # MWhr to MU
        
        print(f"\n--- GROSS GENERATION COMPARISON (MU) ---")
        print(f"Excel MTD (Sum): {mtd_val:.4f}")
        print(f"Excel YTD (Sum): {ytd_val:.4f}")

    # 4. PAF and PLF are usually calculated differently (weighted average or formula based on sums)
    # But let's check the DGR sheet directly for SN 11, 12 if possible.
    ws_dgr = wb['DGR']
    print("\n--- DGR Sheet Extract (Row 10-15) ---")
    for r in range(10, 15):
        row_vals = [ws_dgr.cell(row=r, column=c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")

audit()
