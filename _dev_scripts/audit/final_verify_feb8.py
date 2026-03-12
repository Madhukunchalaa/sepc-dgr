from openpyxl import load_workbook
import datetime

def final_verify():
    excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb['Ops Input']

    # Dates
    target_date = datetime.datetime(2026, 2, 8).date()
    mtd_start = datetime.datetime(2026, 2, 1).date()
    ytd_start = datetime.datetime(2025, 4, 1).date()

    cols = {'target': 318, 'mtd': 311, 'ytd': 5}

    metrics = {
        'Gross Gen (MWhr)': 'gen',
        'Declared capacity': 'dc',
        'Unit On Grid': 'on_grid',
        'Ash Generation': 'ash',
        'DM water Production': 'dm_p',
        'DM Water Consumption total plant': 'dm_c'
    }

    row_map = {}
    for r in range(1, 500):
        found_text = ""
        for c in range(1, 6):
            val = ws.cell(row=r, column=c).value
            if val: found_text += str(val) + " "
        
        if "Gross Generation" in found_text and "Main Meter" in found_text: row_map['gen'] = r
        if "Declared capacity for the day" in found_text: row_map['dc'] = r
        if "hrs" in found_text and "on" in found_text.lower() and "grid" in found_text.lower(): row_map['on_grid'] = r
        if "DM water Production" in found_text: row_map['dm_p'] = r
        if "DM Water Consumption" in found_text and "total plant" in found_text: row_map['dm_c'] = r

    # Ash Generation is usually in DGR sheet or calculated. Let's stick to these for now.

    print(f"{'Metric'.ljust(35)} | {'MTD (Excel)'.ljust(12)} | {'MTD (App)'.ljust(12)} | {'Match'}")
    print("-" * 80)

    # User App Values (Feb 8)
    app_mtd = {
        'gen': 12.9940,
        'dc': 18.0000,
        'on_grid': 3.0000,
        'dm_p': 910.0000,
        'dm_c': 1398.2040
    }

    for key, row in row_map.items():
        mtd_sum = 0
        for c in range(cols['mtd'], cols['target']): # Sum up to Feb 7
            v = ws.cell(row=row, column=c).value
            if v: mtd_sum += float(v)
        
        # Divider for MU
        val = mtd_sum / 1000 if key in ['gen', 'dc'] else mtd_sum
        
        match = "✅" if abs(val - app_mtd[key]) < 0.01 else "❌"
        print(f"{key.ljust(35)} | {val:12.4f} | {app_mtd[key]:12.4f} | {match}")

final_verify()
