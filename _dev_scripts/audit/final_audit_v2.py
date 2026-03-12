from openpyxl import load_workbook
import datetime

def audit_v2():
    excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    ws_dgr = wb['DGR']

    print("--- TAQA DGR Comparison (Excel values) ---")
    
    # SNs to check based on user's snippet
    target_sns = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "17", "19", "25", "28", "29", "32", "38", "39", "44", "49", "59", "61"]
    
    # Find rows for these SNs in DGR sheet (Column 1 has SN)
    print(f"{'SN'.ljust(4)} | {'Particulars'.ljust(35)} | {'Daily'.ljust(10)} | {'MTD'.ljust(10)} | {'YTD'.ljust(12)}")
    print("-" * 80)
    
    for r in range(1, 150):
        sn_val = str(ws_dgr.cell(row=r, column=1).value).strip()
        if sn_val in target_sns:
            particulars = str(ws_dgr.cell(row=r, column=3).value).strip()
            # In 'DGR' sheet, Daily is usually Col 5, MTD is Col 6, YTD is Col 8 (based on SN 7 index I saw earlier)
            # Wait, let's look at SN 7 again: Row 10: [7, None, 'Auxiliary Consumption ', 'MU', val5, val6, val7, val8, val9]
            daily = ws_dgr.cell(row=r, column=5).value
            mtd = ws_dgr.cell(row=r, column=7).value # Based on earlier check, Column 7 was MTD?
            ytd = ws_dgr.cell(row=r, column=8).value # Column 8 was YTD?
            
            # Let's print the whole row first to be sure
            # print(f"DEBUG Row {r}: {[ws_dgr.cell(row=r, column=c).value for c in range(1, 10)]}")
            
            fmt = lambda v: f"{v:10.4f}" if isinstance(v, (int, float)) else str(v).ljust(10)
            print(f"{sn_val.ljust(4)} | {particulars[:35].ljust(35)} | {fmt(daily)} | {fmt(mtd)} | {fmt(ytd)}")

audit_v2()
