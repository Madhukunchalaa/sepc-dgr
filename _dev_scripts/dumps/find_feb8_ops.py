from openpyxl import load_workbook
import datetime

excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
wb = load_workbook(excel_file, read_only=True, data_only=True)
ws = wb['Ops Input']

# Find Feb 8, 2026 column
# Based on earlier check, dates are in header (row 1 or 4?)
# Let's check row 4 and 1
print("--- Ops Input Dates (Row 1) ---")
row1 = [c.value for c in ws[1][:30]]
print(row1)

print("--- Ops Input Dates (Row 4) ---")
row4 = [c.value for c in ws[4][:30]]
print(row4)

target_date = datetime.datetime(2026, 2, 8)
target_col = None

for col in range(1, ws.max_column + 1):
    val = ws.cell(row=4, column=col).value
    if isinstance(val, datetime.datetime) and val.date() == target_date.date():
        target_col = col
        break

if not target_col:
    # Try row 1
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        # If it's a date...
        if isinstance(val, datetime.datetime) and val.date() == target_date.date():
            target_col = col
            break

if target_col:
    print(f"Found Feb 8 at Column {target_col}")
    # Extract some values
    print("Values for Feb 8:")
    for r in range(1, 20):
        param = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=target_col).value
        print(f"Row {r}: {param} = {val}")
else:
    print("Feb 8 not found in Ops Input columns 1 to max.")
