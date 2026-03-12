from openpyxl import load_workbook

excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
wb = load_workbook(excel_file, data_only=True)
ws = wb['DGR']

print("--- Specific Consumption Rows ---")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=60, min_col=1, max_col=10, values_only=True)):
    if row[2] and "Sp " in str(row[2]):
        print(f"Row {i+1}: {row}")
