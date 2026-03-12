import openpyxl
import datetime

file_path = 'TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet = wb['DGR']

print(f"Sheet: {sheet.title}")
for r in range(1, 31):
    # Check if any cell in the first 10 columns of this row is a date
    has_date = False
    for c in range(1, 15):
        val = sheet.cell(row=r, column=c).value
        if isinstance(val, (datetime.datetime, datetime.date)):
            has_date = True
            print(f"Row {r} has date: {val} at Col {c}")
            break
    if not has_date:
        # Also check if it's a string that looks like a date
        for c in range(1, 15):
            val = sheet.cell(row=r, column=c).value
            if val and isinstance(val, str) and (('2025' in val) or ('2026' in val)):
                 print(f"Row {r} potentially has date string: {val} at Col {c}")
                 break
