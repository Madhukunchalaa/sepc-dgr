import openpyxl
import json

excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
try:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    if 'DGR' not in wb.sheetnames:
        print(f"Sheet 'DGR' not found. Available sheets: {wb.sheetnames}")
        exit(1)
        
    ws = wb['DGR']
    print("--- TAQA DGR Excel Structure (First 65 rows) ---")
    
    # Iterate through rows and columns
    for row_idx in range(1, 66):
        row_data = []
        for col_idx in range(1, 11):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is None:
                row_data.append("")
            else:
                row_data.append(str(cell_val).strip())
        
        print(f"Row {row_idx:2}: {' | '.join(row_data)}")

except Exception as e:
    print(f"Error: {e}")
