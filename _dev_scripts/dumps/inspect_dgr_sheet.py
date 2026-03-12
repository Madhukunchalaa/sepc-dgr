from openpyxl import load_workbook
import datetime

excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
wb = load_workbook(excel_file, data_only=True)
ws_dgr = wb['DGR']

# Attempt to find the date cell. Usually it's at the top.
# Let's print first few rows of DGR sheet to find the date cell and the headers.
print("--- DGR Sheet Header ---")
for r in range(1, 15):
    row_vals = [ws_dgr.cell(row=r, column=c).value for c in range(1, 10)]
    print(f"Row {r}: {row_vals}")

# The user's provided values for Feb 8:
app_values = {
    "Gross Generation": {"mtd": 12.9940, "ytd": 1304.1757},
    "APC": {"mtd": 0.0836, "ytd": 0.0808},
    "PAF": {"mtd": 0.3750, "ytd": 0.8026},
    "PLF": {"mtd": 0.2707, "ytd": 0.6922}
}

# Values for Feb 8 in DGR sheet (after setting date if possible)
# If we can't 'set' the date easily, we might need to check 'Ops Input' 
# or look for a summary sheet.
