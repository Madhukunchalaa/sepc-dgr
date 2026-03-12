from openpyxl import load_workbook
import datetime

excel_file = r'c:\Users\IE-Admin\Desktop\dgr\dgr-platform\TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
wb = load_workbook(excel_file, read_only=True, data_only=True)
ws = wb['Ops Input']

target_col = 318  # From find_feb8_ops.py
print(f"--- Parameters for Feb 8 (Col {target_col}) ---")

# Let's extract row 1 to 300 to find all metrics
for r in range(1, 400):
    # Try different columns for the label (usually 1, 2, or 3)
    label1 = ws.cell(row=r, column=1).value
    label2 = ws.cell(row=r, column=2).value
    label3 = ws.cell(row=r, column=3).value
    val = ws.cell(row=r, column=target_col).value
    
    if label1 or label2 or label3:
        # Construct label
        lbl = f"{label1} | {label2} | {label3}".replace("None | ", "").replace(" | None", "").replace("None", "")
        print(f"Row {r}: {lbl} = {val}")

