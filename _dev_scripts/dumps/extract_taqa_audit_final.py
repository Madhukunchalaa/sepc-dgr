import openpyxl
import datetime
import json

file_path = 'TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
target_dates = [
    datetime.date(2025, 7, 20),
    datetime.date(2025, 11, 15),
    datetime.date(2026, 2, 10)
]

def extract_data():
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Extract from DGR (WU) - Calculated KPIs
    sheet_wu = wb['DGR (WU)']
    date_row_wu = 2
    cols_wu = {}
    for c in range(5, 500):
        val = sheet_wu.cell(row=date_row_wu, column=c).value
        if isinstance(val, (datetime.datetime, datetime.date)):
             d = val.date() if isinstance(val, datetime.datetime) else val
             for td in target_dates:
                 if d == td:
                     cols_wu[td.strftime('%Y-%m-%d')] = c
    
    print(f"DGR (WU) Date Columns: {cols_wu}")
    
    # Extract from Ops Input - Raw values
    sheet_ops = wb['Ops Input']
    date_row_ops = 1
    cols_ops = {}
    for c in range(4, 500):
        val = sheet_ops.cell(row=date_row_ops, column=c).value
        if isinstance(val, (datetime.datetime, datetime.date)):
             d = val.date() if isinstance(val, datetime.datetime) else val
             for td in target_dates:
                 if d == td:
                     cols_ops[td.strftime('%Y-%m-%d')] = c
    
    print(f"Ops Input Date Columns: {cols_ops}")
    
    final_results = {}
    for d_str in [d.strftime('%Y-%m-%d') for d in target_dates]:
        if d_str not in cols_wu or d_str not in cols_ops:
            print(f"Missing data for {d_str}")
            continue
            
        day_report = {'kpis': {}, 'raw': {}}
        
        # KPIs from DGR (WU)
        for r in range(4, 80):
            particulars = sheet_wu.cell(row=r, column=3).value
            uom = sheet_wu.cell(row=r, column=4).value
            val = sheet_wu.cell(row=r, column=cols_wu[d_str]).value
            if particulars:
                day_report['kpis'][str(particulars).strip()] = {'uom': str(uom).strip(), 'val': val}
        
        # Raw from Ops Input
        for r in range(4, 180):
            particulars = sheet_ops.cell(row=r, column=2).value
            uom = sheet_ops.cell(row=r, column=3).value
            val = sheet_ops.cell(row=r, column=cols_ops[d_str]).value
            if particulars:
                day_report['raw'][str(particulars).strip()] = {'uom': str(uom).strip(), 'val': val}
                
        final_results[d_str] = day_report
        
    with open('audit_master_data.json', 'w') as f:
        json.dump(final_results, f, default=str, indent=2)
    print("Exported audit_master_data.json")

extract_data()
