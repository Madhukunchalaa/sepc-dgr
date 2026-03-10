import openpyxl
import datetime
import json

file_path = 'TAQA MEIL Neyveli Daily Generation Report Master file 2025-26 R5.xlsx'
# Select 3 stable dates within the range
target_dates = [
    datetime.date(2025, 7, 20),
    datetime.date(2025, 11, 15),
    datetime.date(2026, 2, 10)
]

def get_excel_data(dates):
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb['DGR']
    
    date_row = 9
    date_cols = {}
    
    # Iterate through row 9 to find columns for target dates
    for col in range(5, 500):
        val = sheet.cell(row=date_row, column=col).value
        cell_date = None
        if isinstance(val, datetime.datetime):
            cell_date = val.date()
        elif isinstance(val, datetime.date):
            cell_date = val
            
        if cell_date:
            for td in dates:
                if cell_date == td:
                    date_cols[td.strftime('%Y-%m-%d')] = col
                    break
                    
    print(f"Found date columns: {date_cols}")
    
    if not date_cols:
        # Fallback: maybe dates are strings?
        for col in range(5, 500):
            val = sheet.cell(row=date_row, column=col).value
            if val:
                s_val = str(val).split()[0] # handles 2025-04-01 00:00:00
                for td in dates:
                    if s_val == td.strftime('%Y-%m-%d'):
                        date_cols[td.strftime('%Y-%m-%d')] = col
        print(f"Fallback Search Found: {date_cols}")

    results = {}
    for d_str, col_idx in date_cols.items():
        day_data = {}
        for r in range(12, 160):
            sn = sheet.cell(row=r, column=1).value
            particulars = sheet.cell(row=r, column=2).value
            uom = sheet.cell(row=r, column=3).value
            daily = sheet.cell(row=r, column=col_idx).value
            
            if particulars and str(particulars).strip() and str(particulars) != 'nan':
                key = str(particulars).strip()
                day_data[key] = {
                    'sn': str(sn).strip() if sn else '',
                    'uom': str(uom).strip() if uom else '',
                    'value': daily
                }
        results[d_str] = day_data
        
    return results

results = get_excel_data(target_dates)
with open('audit_excel_values.json', 'w') as f:
    json.dump(results, f, default=str, indent=2)
print("Saved extraction to audit_excel_values.json")
